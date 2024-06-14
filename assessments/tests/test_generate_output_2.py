from unittest.mock import MagicMock, call, patch

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side
from pandas import DataFrame

from assessments.models import ComplianceCriticalityAssessment, Report
from assessments.tests.factories import ComplianceCriticalityAssessmentFactory
from requirements.models import Requirement
from requirements.tests.factories import (
    ComplianceFactory,
    ReferenceFactory,
    ReferencePolicyFactory,
    RequirementFactory,
)

from ..applogic.form_results_applogic import get_form_results
from ..applogic.generate_output_2 import (
    add_bbb_policy,
    add_cloud_control_matrix,
    add_nist_sp800,
    add_references_for_eres_applicability,
    add_references_for_impact_classification,
    df_to_excel,
    generate_output_2,
    insert_data_column,
)


@pytest.fixture
def sample_dataframe():
    # Create a sample dataframe for testing
    data = {"A": [1, 2, 3], "B": ["apple", "banana", "cherry"]}
    return pd.DataFrame(data)


@pytest.fixture
def mock_cca_data():
    return {
        "cca_name": "TestCCA",
        "summary_data": {"some_key": "some_value"},
    }


class RequirementsQuerysetMock:
    def __init__(self, data):
        self.data = data

    def values_list(self, field_name, flat=True):
        return [item[field_name] for item in self.data]


def test_insert_data_column_with_string(sample_dataframe):
    # setup string values
    requirements_data = [
        {"field_name": "value1", "control_id": "A"},
        {"field_name": "value2", "control_id": "B"},
        {"field_name": "value3", "control_id": "C"},
    ]
    requirements_queryset = RequirementsQuerysetMock(requirements_data)
    sample_dataframe.insert(
        0, "Requirement #", requirements_queryset.values_list("control_id", flat=True)
    )

    data = {"A": ["value1"], "B": ["value2"], "C": ["value3"]}
    # Call the function
    insert_data_column(sample_dataframe, 1, "NewColumn", data, "field_name")

    # Assert the changes in the dataframe
    expected_result = pd.DataFrame(
        {
            "Requirement #": ["A", "B", "C"],
            "A": [1, 2, 3],
            "B": ["apple", "banana", "cherry"],
            "NewColumn": ["value1", "value2", "value3"],
        }
    )
    pd.testing.assert_frame_equal(sample_dataframe, expected_result)


def test_insert_data_column_with_list(sample_dataframe):
    # setup list values
    requirements_data = [
        {
            "field_name": ["value1a", "value1b"],
            "control_id": "A",
        },
        {
            "field_name": ["value2a", "value2b"],
            "control_id": "B",
        },
        {
            "field_name": ["value3a", "value3b"],
            "control_id": "C",
        },
    ]
    requirements_queryset = RequirementsQuerysetMock(requirements_data)
    sample_dataframe.insert(
        0, "Requirement #", requirements_queryset.values_list("control_id", flat=True)
    )

    data = {
        "A": ["value1a", "value1b"],
        "B": ["value2a", "value2b"],
        "C": ["value3a", "value3b"],
    }

    # Call the function
    insert_data_column(sample_dataframe, 1, "NewColumn", data, "field_name")

    # Assert the changes in the dataframe
    expected_result = pd.DataFrame(
        {
            "Requirement #": ["A", "B", "C"],
            "A": [1, 2, 3],
            "B": ["apple", "banana", "cherry"],
            "NewColumn": [
                "value1a\r\nvalue1b",
                "value2a\r\nvalue2b",
                "value3a\r\nvalue3b",
            ],
        }
    )
    pd.testing.assert_frame_equal(sample_dataframe, expected_result)


def test_insert_data_column_empty(sample_dataframe):
    requirements_data = [
        {"control_id": "A", "field_name": ""},
        {"control_id": "B", "field_name": ""},
        {"control_id": "C", "field_name": ""},
    ]
    empty_requirements_queryset = RequirementsQuerysetMock(requirements_data)
    sample_dataframe.insert(
        0,
        "Requirement #",
        empty_requirements_queryset.values_list("control_id", flat=True),
    )

    data = {
        "A": [],
        "B": [],
        "C": [],
    }

    # Call the function
    insert_data_column(sample_dataframe, 0, "NewColumn", data, "field_name")

    # Assert the changes in the dataframe
    expected_result = pd.DataFrame(
        {
            "Requirement #": ["A", "B", "C"],
            "A": [1, 2, 3],
            "B": ["apple", "banana", "cherry"],
            "NewColumn": ["N/A", "N/A", "N/A"],
        }
    )
    pd.testing.assert_frame_equal(sample_dataframe, expected_result)


@pytest.mark.django_db
def test_add_cloud_control_matrix():
    df = DataFrame()
    requirements_list = RequirementFactory.create_batch(
        5,
        references=[
            ReferenceFactory(
                policy=ReferencePolicyFactory(header_name="cloud_controls_matrix_v_4.0")
            )
            for _ in range(2)
        ],
    )
    requirement_ids = [req.id for req in requirements_list]
    requirements = Requirement.objects.filter(id__in=requirement_ids)

    df.insert(0, "Requirement #", requirements.values_list("control_id", flat=True))

    with patch(
        "assessments.applogic.generate_output_2.insert_data_column"
    ) as insert_data_column_mock:
        requirements = add_cloud_control_matrix(df, requirements)

        # check that data got annotated properly
        for req in requirements:
            assert hasattr(req, "cloud_controls")

        expected_data_set = {}
        for req in requirements:
            expected_data_set.setdefault(req.control_id, []).append(
                getattr(req, "cloud_controls", "")
            )

        # Assert the behavior of insert_data_column function
        insert_data_column_mock.assert_called_once_with(
            df, 1, "CLOUD CONTROLS MATRIX v 4.0", expected_data_set, "cloud_controls"
        )


@pytest.mark.django_db
def test_add_nist_sp800():
    df = DataFrame()
    # add_nist_sp800 adds a column at index 2,
    # but it can't do so without the previous column index existing, so we add them
    df["EmptyColumn1"] = pd.Series()

    requirements_list = RequirementFactory.create_batch(
        5,
        references=[
            ReferenceFactory(
                policy=ReferencePolicyFactory(header_name="nist_sp800-53_r5")
            )
            for _ in range(2)
        ],
    )
    requirement_ids = [req.id for req in requirements_list]
    requirements = Requirement.objects.filter(id__in=requirement_ids)

    df.insert(0, "Requirement #", requirements.values_list("control_id", flat=True))

    with patch(
        "assessments.applogic.generate_output_2.insert_data_column"
    ) as insert_data_column_mock:
        requirements = add_nist_sp800(df, requirements)

        # check that data got annotated properly
        for req in requirements:
            assert hasattr(req, "nist_sp800")

        expected_data_set = {}
        for req in requirements:
            expected_data_set.setdefault(req.control_id, []).append(
                getattr(req, "nist_sp800", "")
            )

        # Assert the behavior of insert_data_column function
        insert_data_column_mock.assert_called_once_with(
            df, 2, "NIST SP800-53 R5", expected_data_set, "nist_sp800"
        )


@pytest.mark.django_db
def test_add_bbb_policy():
    df = DataFrame()
    # add_bbb_policy adds a column at index 3,
    # but it can't do so without the previous column index existing, so we add them
    df["EmptyColumn1"] = pd.Series()
    df["EmptyColumn2"] = pd.Series()

    requirements_list = RequirementFactory.create_batch(
        5,
        references=[
            ReferenceFactory(
                policy=ReferencePolicyFactory(header_name="bbb_policy__procedure")
            )
            for _ in range(2)
        ],
    )
    requirement_ids = [req.id for req in requirements_list]
    requirements = Requirement.objects.filter(id__in=requirement_ids)

    df.insert(0, "Requirement #", requirements.values_list("control_id", flat=True))

    with patch(
        "assessments.applogic.generate_output_2.insert_data_column"
    ) as insert_data_column_mock:
        requirements = add_bbb_policy(df, requirements)

        # check that data got annotated properly
        for req in requirements:
            assert hasattr(req, "bbb_policy")

        expected_data_set = {}
        for req in requirements:
            expected_data_set.setdefault(req.control_id, []).append(
                getattr(req, "bbb_policy", "")
            )

        # Assert the behavior of insert_data_column function
        insert_data_column_mock.assert_called_once_with(
            df, 3, "bbb Policy/ Procedure", expected_data_set, "bbb_policy"
        )


@pytest.mark.django_db
def test_add_references_for_impact_classification():
    df = pd.DataFrame()
    df["EmptyColumn"] = (
        pd.Series()
    )  # simulate the initial column for requirement control id

    summary = {
        "summary_regulatory_privacy_impact_high_privacy": True,
        "summary_regulatory_privacy_impact_medium_privacy": False,
    }

    reference_policy1 = ReferencePolicyFactory()
    reference_policy2 = ReferencePolicyFactory()
    ComplianceFactory(
        header_name="high_privacy", reference_policies=[reference_policy1]
    )
    ComplianceFactory(
        header_name="medium_privacy", reference_policies=[reference_policy2]
    )

    req1 = RequirementFactory(references=[ReferenceFactory(policy=reference_policy1)])
    req2 = RequirementFactory(references=[ReferenceFactory(policy=reference_policy2)])
    requirements = Requirement.objects.filter(id__in=[req1.id, req2.id])

    with patch(
        "assessments.applogic.generate_output_2.insert_data_column"
    ) as insert_data_column_mock:
        requirements = add_references_for_impact_classification(
            df, summary, requirements
        )

        # check that data got annotated properly
        data_set = []
        for req in requirements:
            for index, policy in enumerate([reference_policy1, reference_policy2]):
                field_name = f"impact_reference_{policy.id}"
                assert hasattr(req, field_name)

                data_dict = {}
                for req in requirements:
                    data_dict.setdefault(req.control_id, []).append(
                        getattr(req, field_name, "")
                    )

                data_set.append(data_dict)

        # check that all references got added
        expected_calls = [
            call(
                df,
                1,
                reference_policy1.name,
                data_set[0],
                f"impact_reference_{reference_policy1.id}",
            ),
            call(
                df,
                1,
                reference_policy2.name,
                data_set[1],
                f"impact_reference_{reference_policy2.id}",
            ),
        ]

        assert insert_data_column_mock.call_count == 2
        assert repr(insert_data_column_mock.call_args_list[0]) == repr(
            expected_calls[0]
        )
        assert repr(insert_data_column_mock.call_args_list[1]) == repr(
            expected_calls[1]
        )


@pytest.mark.django_db
def test_add_references_for_eres_applicability():
    df = pd.DataFrame()
    df["EmptyColumn"] = (
        pd.Series()
    )  # simulate the initial column for requirement control id

    summary = {
        "summary_regulatory_gxp_eres_er": True,
        "summary_regulatory_gxp_eres_es": False,
    }

    reference_policy1 = ReferencePolicyFactory()
    reference_policy2 = ReferencePolicyFactory()
    ComplianceFactory(header_name="impact", reference_policies=[reference_policy1])
    ComplianceFactory(header_name="no_impact", reference_policies=[reference_policy2])

    req1 = RequirementFactory(references=[ReferenceFactory(policy=reference_policy1)])
    req2 = RequirementFactory(references=[ReferenceFactory(policy=reference_policy2)])
    requirements = Requirement.objects.filter(id__in=[req1.id, req2.id])

    with patch(
        "assessments.applogic.generate_output_2.insert_data_column"
    ) as insert_data_column_mock:
        requirements = add_references_for_eres_applicability(df, summary, requirements)

        # check that data got annotated properly
        data_set = []
        for req in requirements:
            for index, policy in enumerate([reference_policy1]):
                field_name = f"eres_applicability_{policy.id}"
                assert hasattr(req, field_name)

                data_dict = {}
                for req in requirements:
                    data_dict.setdefault(req.control_id, []).append(
                        getattr(req, field_name, "")
                    )

                data_set.append(data_dict)

        # check that all references got added
        expected_calls = [
            call(
                df,
                1,
                reference_policy1.name,
                data_set[0],
                f"eres_applicability_{reference_policy1.id}",
            ),
        ]
        assert insert_data_column_mock.call_count == 1
        assert repr(insert_data_column_mock.call_args_list[0]) == repr(
            expected_calls[0]
        )


def test_df_to_excel(sample_dataframe):
    # Create a workbook and call the function
    wb = Workbook()
    ws = wb.active
    df_to_excel(sample_dataframe, ws)

    # Check formatting for the first cell in the worksheet
    first_cell = ws.cell(row=1, column=1)

    # Check alignment
    assert first_cell.alignment.wrap_text is True
    assert first_cell.alignment.horizontal == "left"
    assert first_cell.alignment.vertical == "top"

    # Check font
    assert first_cell.font.name == "Calibri Light"
    assert first_cell.font.size == 11

    # Check border
    assert first_cell.border.left.style == "thin"
    assert first_cell.border.right.style == "thin"
    assert first_cell.border.top.style == "thin"
    assert first_cell.border.bottom.style == "thin"

    # Check background color
    expected_color = PatternFill(
        start_color="D0CECE", end_color="D0CECE", fill_type="solid"
    )
    assert repr(first_cell.fill.start_color) == repr(expected_color.start_color)
    assert repr(first_cell.fill.end_color) == repr(expected_color.end_color)
    assert repr(first_cell.fill.fill_type) == repr(expected_color.fill_type)


@pytest.mark.django_db
@patch("assessments.applogic.generate_output_2.add_cloud_control_matrix")
@patch("assessments.applogic.generate_output_2.add_nist_sp800")
@patch("assessments.applogic.generate_output_2.add_bbb_policy")
@patch(
    "assessments.applogic.generate_output_2.add_references_for_impact_classification"
)
@patch("assessments.applogic.generate_output_2.add_references_for_eres_applicability")
@patch("assessments.applogic.generate_output_2.df_to_excel")
def test_generate_output_2(*args):
    (
        mock_df_to_excel,
        mock_references_for_eres_applicability,
        mock_references_for_impact_classification,
        mock_bbb_policy,
        mock_nist_sp800,
        mock_cloud_control_matrix,
    ) = args

    # Call the function with the mocked objects
    cca = ComplianceCriticalityAssessmentFactory(
        hosting_and_type=ComplianceCriticalityAssessment.HostingAndType.IAAS
    )
    requirements = get_form_results(cca)
    generate_output_2(cca, requirements)

    assert mock_df_to_excel.called_once()
    assert mock_references_for_eres_applicability.called_once()
    assert mock_references_for_impact_classification.called_once()
    assert mock_bbb_policy.called_once()
    assert mock_nist_sp800.called_once()
    assert mock_cloud_control_matrix.called_once()

    report = Report.objects.filter(cca=cca)
    assert report.count() == 1
    assert report[0].output_xlsm_2 is not None
