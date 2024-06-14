import os
import re
import tempfile

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from django.conf import settings
from django.core.files import File


def get_assessment_models():
    # pragma: no cover
    from .. import models as assessment_models

    return assessment_models


def generate_output_1(cca, requirements):
    # MS excel can't open the xlsm file when we create a new file directly.
    # A workaround is to open an existing empty xlsm file and saving that as our new output file.
    template = os.path.join(settings.STATICFILES_DIRS[0], "data/output_template.xlsm")
    wb = openpyxl.load_workbook(filename=template, read_only=False, keep_vba=True)

    df = pd.DataFrame()

    # add the requirement control ID as the first column
    df.insert(0, "Requirement #", requirements.values_list("control_id", flat=True))
    df.insert(
        1,
        "Requirement Statement",
        requirements.values_list("requirement_statement", flat=True),
    )
    df.insert(2, "Comments", "")
    df.insert(3, "", pd.Series())
    df.insert(
        4,
        "bbb Common Solution",
        requirements.values_list("bbb_common_solution", flat=True),
    )
    df.insert(5, "Test Guidance", requirements.values_list("test_guidance", flat=True))

    df_to_excel(df, wb.active)
    insert_categories(wb.active)

    # Create a temporary Excel file on the local file system
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_excel_file:
        wb.save(temp_excel_file.name)

    # Open the temporary file for reading and save it to the model's FileField
    with open(temp_excel_file.name, "rb") as excel_file:
        assessment_models = get_assessment_models()
        report, created = assessment_models.Report.objects.get_or_create(cca=cca)
        report.output_xlsm_1.save(f"output1.xlsm", File(excel_file))

    # Delete the temporary file
    os.unlink(temp_excel_file.name)


def df_to_excel(df, ws):
    """Write DataFrame df to openpyxl worksheet ws"""

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # format cells
            cell.alignment = Alignment(
                wrap_text=True,
                horizontal="left",
                vertical="bottom" if c_idx == 2 or c_idx == 6 else "top",
            )
            cell.font = Font(name="Calibri Light", size=11)
            cell.border = cell.border.copy(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    # Autofit columns
    for col in ws.columns:
        max_length = 0
        column = [col[0].column_letter] + [cell for cell in col]
        for cell in column:
            try:
                value = str(cell.value)
                lines = value.split("\n")
                max_line_length = max(len(line) for line in lines)
                if max_line_length > max_length:
                    max_length = max_line_length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width


def get_category(requirement_code):
    """
    This function takes a Requirement control ID that has a similar pattern to another control ID and uses regex to extract
    the to extract the category code and return the category name.
    """
    category_code = re.search(r"_(.*?)-", requirement_code).group(1)
    match category_code:
        case "AM":
            return "Access Management"
        case "CM":
            return "Capability Management"
        case "DG":
            return "Data Governance"
        case "DP":
            return "Data Protection"
        case "DPri":
            return "Data Privacy"
        case "ERES":
            return "Electronic Signatures, Digital Signatures and Electronic Records"
        case "IM":
            return "Incident Management"
        case "IP":
            return "Infrastructure Protection"
        case "LM":
            return "Logging & Monitoring"
        case "PS":
            return "Physical Security"
        case "RM":
            return "Risk Management"
        case "SD":
            return "Secure Development / SDLC"
        case "SM":
            return "Supplier Management"
        case "TA":
            return "Training & Awareness"
        case _:
            return ""


def insert_categories(ws):
    current_category = None
    rows_to_insert = []

    for row_index in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_index, column=1).value
        if cell_value:
            category = get_category(cell_value)
            if category != current_category:
                current_category = category
                title_row = (f"{current_category}",)
                rows_to_insert.append((row_index, title_row))

    # Number of Category Titles
    inserted_rows = len(rows_to_insert) - 1
    category_collor_fill = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )
    category_alignment = Alignment(horizontal="left")
    # Insert the title rows in reverse order to avoid shifting issues
    for row_index, title_row in reversed(rows_to_insert):
        ws.insert_rows(row_index, amount=1)
        for col_index, value in enumerate(title_row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if row_index == 2:
                cell = ws.cell(
                    row=1,
                    column=col_index,
                )
                # Swap the first row with the second row
                for col_index in range(1, ws.max_column + 1):
                    first_row_value = ws.cell(row=1, column=col_index).value
                    second_row_value = ws.cell(row=2, column=col_index).value
                    ws.cell(
                        row=1,
                        column=col_index,
                        value=second_row_value if col_index < 2 else "",
                    )
                    ws.cell(row=2, column=col_index, value=first_row_value)
                ws.merge_cells(
                    start_row=1,
                    start_column=1,
                    end_row=1,
                    end_column=3,
                )
            else:
                # Merge cells across three columns
                ws.merge_cells(
                    start_row=row_index + inserted_rows,
                    start_column=col_index,
                    end_row=row_index + inserted_rows,
                    end_column=col_index + 2,
                )
            # Apply fill color
            cell.fill = category_collor_fill
            # Set alignment to left
            cell.alignment = category_alignment
            inserted_rows -= 1

    # Apply borders to all cells in the second row
    for col_index in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col_index)
        cell.border = cell.border.copy(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
