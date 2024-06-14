import os
import tempfile

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from django.conf import settings
from django.contrib.postgres.aggregates import ArrayAgg
from django.core.files import File
from django.db.models import Case, CharField, F, Q, Value, When

from assessments.applogic import form_results_applogic
from requirements.models import Compliance, ReferencePolicy


def get_assessment_models():
    # pragma: no cover
    from .. import models as assessment_models

    return assessment_models


def transform_value(x):
    """Transforms the value into the expected output for excel

    If the value is a list, convert the list into a new line (/r/n) separated string.
        And strip out any empty values.
    If the value is a string, return the value.
        But if it is an empty string, return 'N/A'.
    If there value is any other data type, return the value itself.
    """
    if isinstance(x, list):
        return (
            "\r\n".join(filter(lambda s: s.strip(), map(str, x)))
            if any(s.strip() for s in x)
            else "N/A"
        )
    elif isinstance(x, str):
        return x if x else "N/A"
    else:
        return x


def insert_data_column(df, column_index, column_name, data, field_name):
    """Inserts the data into the dataframe

    Args:
        df (pandas.dataframe): Pandas Dataframe object that contains all our data to be exported
        column_index (int): O-based Index column position to insert the data
        column_name (str): Name of the column
        requirements (obj): Requirements queryest to pull the data from
        field_name (str): Attribute on the Requirements queryset to pull data from
    """
    # insert the annotated data into the dataframe
    # df.insert(
    #     column_index, column_name, requirements.values_list(field_name, flat=True)
    # )
    df[column_name] = df["Requirement #"].map(data)

    # format the cell and convert the list from values_list into a string with each item in the list is placed in a new line
    df[column_name] = df[column_name].apply(transform_value)


def add_cloud_control_matrix(df, requirements):
    """Add the CLOUD CONTROLS MATRIX v 4.0 into the dataframe

    This is a fixed column where row values are directly copied for each requirement.
    Should be added to column B (index 1)
    """
    column_name = "CLOUD CONTROLS MATRIX v 4.0"
    policy_header_name = "cloud_controls_matrix_v_4.0"
    column_index = 1
    field_name = "cloud_controls"

    annotations = {
        # field_name: ArrayAgg(
        #     Case(
        #         When(
        #             references__policy__header_name=policy_header_name,
        #             then=F("references__identifier"),
        #         ),
        #         default=Value(""),
        #         output_field=CharField(),
        #     ),
        #     distinct=True,
        # ),
        field_name: Case(
            When(
                references__policy__header_name=policy_header_name,
                then=F("references__identifier"),
            ),
            default=Value(""),
            output_field=CharField(),
        ),
    }

    requirements = requirements.annotate(**annotations)

    data_dict = {}
    for req in requirements:
        data_dict.setdefault(req.control_id, []).append(getattr(req, field_name, ""))

    insert_data_column(df, column_index, column_name, data_dict, field_name)

    return requirements


def add_nist_sp800(df, requirements):
    """Add the NIST SP800-53 R5 into the dataframe

    This is a fixed column where row values are directly copied for each requirement.
    Should be added to column C (index 2)
    """
    column_name = "NIST SP800-53 R5"
    policy_header_name = "nist_sp800-53_r5"
    column_index = 2
    field_name = "nist_sp800"

    annotations = {
        # field_name: ArrayAgg(
        #     Case(
        #         When(
        #             references__policy__header_name=policy_header_name,
        #             then=F("references__identifier"),
        #         ),
        #         default=Value(""),
        #         output_field=CharField(),
        #     ),
        #     distinct=True,
        # ),
        field_name: Case(
            When(
                references__policy__header_name=policy_header_name,
                then=F("references__identifier"),
            ),
            default=Value(""),
            output_field=CharField(),
        ),
    }
    requirements = requirements.annotate(**annotations)

    data_dict = {}
    for req in requirements:
        data_dict.setdefault(req.control_id, []).append(getattr(req, field_name, ""))

    insert_data_column(df, column_index, column_name, data_dict, field_name)

    return requirements


def add_bbb_policy(df, requirements):
    """Add the bbb Policy/ Procedure into the dataframe

    This is a fixed column where row values are directly copied for each requirement.
    Should be added to column D (index 3)
    """
    column_name = "bbb Policy/ Procedure"
    policy_header_name = "bbb_policy__procedure"
    column_index = 3
    field_name = "bbb_policy"

    annotations = {
        # field_name: ArrayAgg(
        #     Case(
        #         When(
        #             references__policy__header_name=policy_header_name,
        #             then=F("references__identifier"),
        #         ),
        #         default=Value(""),
        #         output_field=CharField(),
        #     ),
        #     distinct=True,
        # ),
        field_name: Case(
            When(
                references__policy__header_name=policy_header_name,
                then=F("references__identifier"),
            ),
            default=Value(""),
            output_field=CharField(),
        )
    }
    requirements = requirements.annotate(**annotations)

    data_dict = {}
    for req in requirements:
        data_dict.setdefault(req.control_id, []).append(getattr(req, field_name, ""))

    insert_data_column(df, column_index, column_name, data_dict, field_name)

    return requirements


def add_references_for_impact_classification(df, summary, requirements):
    """Adds a column for each Reference associated with impact classification

    References (columns) should only be added if the CCA is of impact 7 or 8 (high privacy or medium privacy)
    """
    is_cca_high_or_medium_privacy = (
        summary["summary_regulatory_privacy_impact_high_privacy"]
        or summary["summary_regulatory_privacy_impact_medium_privacy"]
    )

    # only add the columns if cca is high or medium privacy impact
    if is_cca_high_or_medium_privacy:
        # get references policies to add (columns)
        compliances = Compliance.objects.filter(
            Q(header_name="high_privacy") | Q(header_name="medium_privacy")
        )
        reference_policy_ids = compliances.values_list(
            "reference_policies__id", flat=True
        ).distinct()

        for reference_policy in ReferencePolicy.objects.filter(
            id__in=reference_policy_ids
        ):
            column_name = reference_policy.name
            field_name = f"impact_reference_{reference_policy.id}"

            # last column index + 1, basically add a new column
            column_index = df.columns.get_loc(df.columns[-1]) + 1

            annotations = {
                # field_name: ArrayAgg(
                #     Case(
                #         When(
                #             references__policy__id=reference_policy.id,
                #             then=F("references__identifier"),
                #         ),
                #         default=Value(""),
                #         output_field=CharField(),
                #     ),
                #     distinct=True,
                # ),
                field_name: Case(
                    When(
                        references__policy__id=reference_policy.id,
                        then=F("references__identifier"),
                    ),
                    default=Value(""),
                    output_field=CharField(),
                )
            }

            requirements = requirements.annotate(**annotations)

            data_dict = {}
            for req in requirements:
                data_dict.setdefault(req.control_id, []).append(
                    getattr(req, field_name, "")
                )

            insert_data_column(df, column_index, column_name, data_dict, field_name)

    return requirements


def add_references_for_eres_applicability(df, summary, requirements):
    """Adds a column for each Reference associated with ERES applicability

    References (columns) should only be added if the CCA is of (ERES) Impact or No Impact
    """
    has_eres_impact = (
        summary["summary_regulatory_gxp_eres_er"]
        or summary["summary_regulatory_gxp_eres_es"]
    )

    if has_eres_impact:
        compliances = Compliance.objects.filter(header_name="impact")
    else:
        compliances = Compliance.objects.filter(header_name="no_impact")

    reference_policy_ids = compliances.values_list(
        "reference_policies__id", flat=True
    ).distinct()

    for reference_policy in ReferencePolicy.objects.filter(id__in=reference_policy_ids):
        column_name = reference_policy.name
        field_name = f"eres_applicability_{reference_policy.id}"

        # last column index + 1, basically add a new column
        column_index = df.columns.get_loc(df.columns[-1]) + 1

        annotations = {
            # field_name: ArrayAgg(
            #     Case(
            #         When(
            #             references__policy__id=reference_policy.id,
            #             then=F("references__identifier"),
            #         ),
            #         default=Value(""),
            #         output_field=CharField(),
            #     ),
            #     distinct=True,
            # ),
            field_name: Case(
                When(
                    references__policy__id=reference_policy.id,
                    then=F("references__identifier"),
                ),
                default=Value(""),
                output_field=CharField(),
            ),
        }

        requirements = requirements.annotate(**annotations)

        data_dict = {}
        for req in requirements:
            data_dict.setdefault(req.control_id, []).append(
                getattr(req, field_name, "")
            )

        insert_data_column(df, column_index, column_name, data_dict, field_name)

    return requirements


def df_to_excel(df, ws):
    """Write DataFrame df to openpyxl worksheet ws"""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # format cells
            cell.alignment = Alignment(
                wrap_text=True, horizontal="left", vertical="top"
            )
            cell.font = Font(name="Calibri Light", size=11)
            cell.border = cell.border.copy(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    # Autofit columns
    for col in ws.columns:
        max_length = 0
        column = [col[0].column_letter] + [cell for cell in col]
        for cell in column:
            try:
                value = str(cell.value)
                lines = value.split("\n")
                max_line_length = max(len(line) for line in lines)
                if max_line_length > max_length:
                    max_length = max_line_length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Apply background color to the first row
    for cell in ws[1]:
        cell.fill = PatternFill(
            start_color="D0CECE", end_color="D0CECE", fill_type="solid"
        )


def generate_output_2(cca, requirements):
    # MS excel can't open the xlsm file when we create a new file directly.
    # A workaround is to open an existing empty xlsm file and saving that as our new output file.
    template = os.path.join(settings.STATICFILES_DIRS[0], "data/output_template.xlsm")
    wb = openpyxl.load_workbook(filename=template, read_only=False, keep_vba=True)
    ws = wb.active
    ws.title = "Output 2"

    df = pd.DataFrame()

    # add the requirement control ID as the first column
    df.insert(0, "Requirement #", requirements.values_list("control_id", flat=True))

    # add static columns
    add_cloud_control_matrix(df, requirements)
    add_nist_sp800(df, requirements)
    add_bbb_policy(df, requirements)

    # add summary based dynamic columns
    add_references_for_impact_classification(df, cca.summary, requirements)
    add_references_for_eres_applicability(df, cca.summary, requirements)

    # add static requirement data columns at the end
    last_column_index = df.columns.get_loc(df.columns[-1])
    df.insert(
        last_column_index + 1,
        "bbb Common Solution",
        requirements.values_list("bbb_common_solution", flat=True),
    )
    df.insert(
        last_column_index + 2,
        "Test Guidance",
        requirements.values_list("test_guidance", flat=True),
    )

    # Write the dataframe into the workbook active sheet
    df_to_excel(df, ws)

    # Create a temporary Excel file on the local file system
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_excel_file:
        wb.save(temp_excel_file.name)

    # Open the temporary file for reading and save it to the model's FileField
    with open(temp_excel_file.name, "rb") as excel_file:
        assessment_models = get_assessment_models()
        report, created = assessment_models.Report.objects.get_or_create(cca=cca)
        report.output_xlsm_2.save(f"output2.xlsm", File(excel_file))

    # Delete the temporary file
    os.unlink(temp_excel_file.name)
